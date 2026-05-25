import csv
import io
import uuid
from datetime import datetime, timedelta

import openpyxl
from openpyxl.styles import Alignment, Font
from weasyprint import HTML


class ReportExporter:
    """Handles generation of report files in Excel, CSV, and PDF formats."""

    def __init__(self, supabase_client):
        self.supabase = supabase_client
        self._jobs: dict[str, dict] = {}

    def generate_excel(
        self,
        columns: list[dict],
        rows: list[dict],
        title: str = "Report",
    ) -> bytes:
        """Generate an Excel workbook from column/row data with RTL support."""
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = title[:31]

        # Write headers
        headers = [col.get("label", col.get("key", "")) for col in columns]
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="right", readingOrder=2)

        # Write rows
        for row in rows:
            values = []
            for col in columns:
                key = col["key"]
                val = row.get(key, "")
                if val is None:
                    val = ""
                values.append(val)
            ws.append(values)

        # Adjust column widths
        for idx, col in enumerate(columns, 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(idx)].width = max(
                15, len(col.get("label", "")) + 5
            )

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.read()

    def generate_csv(
        self,
        columns: list[dict],
        rows: list[dict],
    ) -> bytes:
        """Generate a CSV file from column/row data."""
        buffer = io.StringIO()
        writer = csv.writer(buffer)
        headers = [col.get("label", col.get("key", "")) for col in columns]
        writer.writerow(headers)
        for row in rows:
            values = [row.get(col["key"], "") for col in columns]
            writer.writerow(values)
        return buffer.getvalue().encode("utf-8-sig")

    def generate_pdf(
        self,
        columns: list[dict],
        rows: list[dict],
        title: str = "Report",
    ) -> bytes:
        """Generate a PDF from column/row data using WeasyPrint with RTL CSS."""
        headers = [col.get("label", col.get("key", "")) for col in columns]
        html_rows = []
        for row in rows:
            cells = [str(row.get(col["key"], "")) for col in columns]
            html_rows.append("<tr>" + "".join(f"<td>{c}</td>" for c in cells) + "</tr>")

        html_content = f"""
        <!DOCTYPE html>
        <html dir="rtl">
        <head>
            <meta charset="UTF-8">
            <style>
                body {{ font-family: 'DejaVu Sans', sans-serif; margin: 20px; }}
                table {{ width: 100%; border-collapse: collapse; }}
                th, td {{ border: 1px solid #ccc; padding: 8px; text-align: right; }}
                th {{ background-color: #f0f0f0; font-weight: bold; }}
                h1 {{ text-align: center; }}
            </style>
        </head>
        <body>
            <h1>{title}</h1>
            <table>
                <thead>
                    <tr>{"".join(f"<th>{h}</th>" for h in headers)}</tr>
                </thead>
                <tbody>
                    {"".join(html_rows)}
                </tbody>
            </table>
        </body>
        </html>
        """
        return HTML(string=html_content).write_pdf()

    def generate_file(
        self,
        columns: list[dict],
        rows: list[dict],
        export_format: str,
        title: str = "Report",
    ) -> bytes:
        """Dispatch to the correct generator based on format."""
        if export_format == "xlsx":
            return self.generate_excel(columns, rows, title)
        elif export_format == "csv":
            return self.generate_csv(columns, rows)
        elif export_format == "pdf":
            return self.generate_pdf(columns, rows, title)
        else:
            raise ValueError(f"Unsupported export format: {export_format}")

    async def upload_to_storage(
        self,
        file_bytes: bytes,
        file_name: str,
        org_id: str,
    ) -> str:
        """Upload generated report to Supabase Storage and return the file path."""
        bucket = "report-archives"
        path = f"{org_id}/{datetime.utcnow().strftime('%Y/%m')}/{uuid.uuid4()}_{file_name}"
        try:
            # Ensure bucket exists
            buckets = self.supabase.storage.list_buckets()
            bucket_names = [b.name for b in buckets]
            if bucket not in bucket_names:
                self.supabase.storage.create_bucket(bucket, options={"public": False})
        except Exception:
            pass  # Bucket may already exist or permissions restrict listing

        self.supabase.storage.from_(bucket).upload(path, file_bytes)
        return path

    def create_job(self, filters: dict, export_format: str) -> str:
        """Create an async export job and return its ID."""
        job_id = str(uuid.uuid4())
        self._jobs[job_id] = {
            "job_id": job_id,
            "status": "generating",
            "progress_pct": 0,
            "download_url": None,
            "error": None,
            "filters": filters,
            "format": export_format,
            "created_at": datetime.utcnow().isoformat(),
        }
        return job_id

    def update_job(self, job_id: str, status: str, progress_pct: int = 0, download_url: str | None = None, error: str | None = None):
        """Update the status of an async export job."""
        if job_id in self._jobs:
            self._jobs[job_id]["status"] = status
            self._jobs[job_id]["progress_pct"] = progress_pct
            if download_url is not None:
                self._jobs[job_id]["download_url"] = download_url
            if error is not None:
                self._jobs[job_id]["error"] = error

    def get_job(self, job_id: str) -> dict | None:
        """Get the current status of an async export job."""
        return self._jobs.get(job_id)

    def cleanup_old_jobs(self, max_age_hours: int = 24):
        """Remove jobs older than max_age_hours."""
        cutoff = datetime.utcnow() - timedelta(hours=max_age_hours)
        to_remove = []
        for job_id, job in self._jobs.items():
            created = datetime.fromisoformat(job["created_at"])
            if created < cutoff:
                to_remove.append(job_id)
        for job_id in to_remove:
            del self._jobs[job_id]
